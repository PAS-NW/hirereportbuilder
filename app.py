import io
import re
import zipfile
import tempfile
from copy import copy
from datetime import datetime, date
from pathlib import Path
from html import escape

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
except Exception:
    colors = None

PAS_YELLOW = "#FFD400"
PAS_BLACK = "#0A0A0A"
PAS_DARK = "#171717"
PAS_GREY = "#F4F4F4"
APP_VERSION = "v1.0 Prototype Build"
TEMPLATE_PATHS = [
    Path("assets/Hire Report Template.xlsx"),
    Path("Hire Report Template.xlsx"),
]

st.set_page_config(page_title="PAS Hire Report Builder", page_icon="pas_logo.png", layout="wide")

st.markdown(
    f"""
    <style>
    .stApp {{ background: #f5f5f5; color: #0A0A0A; }}
    section[data-testid="stSidebar"] {{
        background: {PAS_BLACK};
        color: white;
        padding-top: 1.45rem;
    }}
    section[data-testid="stSidebar"] * {{ color: white; }}
    section[data-testid="stSidebar"] img {{
        margin-top: 0.15rem;
        border-radius: 14px;
    }}
    .block-container {{
        padding-top: 1.4rem;
        padding-bottom: 2rem;
        max-width: 1500px;
    }}

    .pas-hero {{
        background: linear-gradient(135deg, {PAS_BLACK} 0%, #202020 70%, #7a6900 135%);
        border-radius: 18px;
        padding: 24px 28px;
        margin-bottom: 18px;
        box-shadow: 0 8px 25px rgba(0,0,0,0.12);
    }}
    .pas-title {{
        color: white;
        font-size: 32px;
        font-weight: 900;
        margin: 0;
        letter-spacing: -0.03em;
    }}
    .pas-subtitle {{
        color: {PAS_YELLOW};
        font-size: 14px;
        margin-top: 4px;
        font-weight: 800;
    }}

    .kpi-card {{
        background: white;
        border-radius: 18px;
        padding: 18px 20px;
        border: 1px solid #e8e8e8;
        box-shadow: 0 3px 12px rgba(0,0,0,0.05);
        min-height: 112px;
    }}
    .kpi-label {{
        color: #111;
        font-size: 14px;
        font-weight: 800;
        margin-bottom: 8px;
    }}
    .kpi-value {{
        color: {PAS_YELLOW};
        font-size: 36px;
        font-weight: 950;
        line-height: 1.05;
        text-shadow: 0 1px 0 #111;
    }}
    .kpi-sub {{
        color: #222;
        font-size: 13px;
        margin-top: 6px;
    }}

    .stButton > button, .stDownloadButton > button {{
        background: {PAS_YELLOW} !important;
        color: {PAS_BLACK} !important;
        border: 1px solid {PAS_BLACK} !important;
        border-radius: 12px !important;
        font-weight: 900 !important;
    }}

    .stCaption, div[data-testid="stCaptionContainer"], .stMarkdown p, .stInfo {{
        color: #0A0A0A !important;
    }}

    .pas-results-title {{
        color: #0A0A0A;
        font-size: 26px;
        font-weight: 950;
        margin: 22px 0 8px 0;
    }}
    .pas-unmatched-pill {{
        display: inline-flex;
        align-items: center;
        gap: 8px;
        background: {PAS_YELLOW};
        color: {PAS_BLACK};
        border: 1px solid #111;
        border-radius: 14px 14px 0 0;
        padding: 11px 18px;
        font-weight: 950;
        box-shadow: 0 3px 10px rgba(0,0,0,0.08);
        margin-top: 4px;
    }}

    .pas-table-wrap {{
        background: white;
        border: 1px solid #d9d9d9;
        border-radius: 0 16px 16px 16px;
        overflow: auto;
        box-shadow: 0 4px 18px rgba(0,0,0,0.07);
        margin-bottom: 18px;
    }}
    table.pas-table {{
        width: 100%;
        border-collapse: collapse;
        font-size: 13px;
        color: #0A0A0A;
        background: white;
    }}
    table.pas-table thead th {{
        background: {PAS_YELLOW};
        color: {PAS_BLACK};
        font-weight: 950;
        text-align: left;
        padding: 11px 12px;
        border: 1px solid #c7a900;
        white-space: nowrap;
    }}
    table.pas-table tbody td {{
        background: white;
        color: #0A0A0A;
        padding: 9px 12px;
        border: 1px solid #e3e3e3;
        vertical-align: top;
    }}
    table.pas-table tbody tr:nth-child(even) td {{ background: #fbfbfb; }}
    table.pas-table a {{ color: #006fd6 !important; font-weight: 800; text-decoration: none; }}
    .pas-note {{ color: #0A0A0A; font-size: 13px; margin: 8px 0 16px 0; }}
    .pas-support {{ color: #0A0A0A; font-size: 14px; margin: 16px 0; }}

    div[data-testid="stAlert"], div[data-testid="stAlert"] * {{ color: #0A0A0A !important; }}
    div[data-testid="stAlert"] {{ border: 1px solid #e2ba00 !important; }}
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown(
    f"""
    <style>
    section[data-testid="stSidebar"] .stMarkdown p,
    section[data-testid="stSidebar"] .stMarkdown li,
    section[data-testid="stSidebar"] .stMarkdown h1,
    section[data-testid="stSidebar"] .stMarkdown h2,
    section[data-testid="stSidebar"] .stMarkdown h3,
    section[data-testid="stSidebar"] .stMarkdown strong,
    section[data-testid="stSidebar"] .stMarkdown span {{ color: #ffffff !important; }}

    div[data-testid="stFileUploader"] svg,
    div[data-testid="stFileUploader"] button svg,
    div[data-testid="stFileUploader"] [data-testid="stIconMaterial"] {{
        color: #FFD400 !important;
        fill: #FFD400 !important;
        stroke: #FFD400 !important;
    }}
    div[data-testid="stFileUploader"] section {{
        background: #24242d !important;
        border: 1px solid #30303a !important;
        border-radius: 12px !important;
    }}
    div[data-testid="stFileUploader"] button {{
        color: white !important;
        border-color: #454552 !important;
        background: #111217 !important;
    }}
    .pas-table-wrap {{ max-height: 510px !important; overflow-y: auto !important; overflow-x: auto !important; }}
    .pas-table-wrap thead th {{ position: sticky; top: 0; z-index: 2; }}
    .pas-note, .pas-support, .pas-support * {{ color: #0A0A0A !important; }}
    .pas-sidebar-title {{ color:#fff; font-size:18px; font-weight:950; line-height:1.15; text-align:center; margin: 20px 0 8px; }}
    .pas-yellow-line {{ width:72px; height:4px; background:{PAS_YELLOW}; border-radius:99px; margin: 0 auto 22px; }}
    .pas-sidebar-copy {{ color:#fff !important; font-size:14px; line-height:1.52; font-weight:650; margin-bottom:24px; }}
    .pas-sidebar-rule {{ border-top:1px solid rgba(255,255,255,.22); margin:22px 0; }}
    .pas-sidebar-heading {{ color:{PAS_YELLOW}; font-size:19px; font-weight:950; margin: 0 0 16px; }}
    .pas-nav-row {{ display:grid; grid-template-columns: 26px 1fr; gap:10px; align-items:start; margin: 15px 0; color:#fff; font-weight:750; line-height:1.25; font-size:14px; }}
    .pas-nav-icon svg {{ width:21px; height:21px; stroke:{PAS_YELLOW}; stroke-width:2.4; fill:none; stroke-linecap:round; stroke-linejoin:round; }}
    .pas-sidebar-footer {{ color:#fff; font-size:12px; font-weight:800; margin-top:28px; }}

    .pas-hero {{ display:flex; align-items:center; gap:16px; background: linear-gradient(100deg, #08090b 0%, #151718 70%, #c9aa00 130%) !important; border-radius: 16px !important; padding: 12px 22px !important; margin: 0 0 18px 0 !important; box-shadow: 0 9px 25px rgba(0,0,0,.13) !important; min-height:60px; }}
    .pas-hero-logo {{ width:37px; height:37px; border-radius:7px; background:{PAS_YELLOW}; color:#000; display:inline-flex; align-items:center; justify-content:center; font-weight:950; font-size:14px; letter-spacing:-1px; }}
    .pas-hero-text {{ color:#fff; font-size:18px; font-weight:950; letter-spacing:-.02em; }}
    .pas-hero-dot {{ color:#fff; opacity:.8; margin: 0 7px; }}
    .pas-hero-version {{ color:{PAS_YELLOW}; font-weight:950; }}

    .pas-upload-card {{ background:#fff; border:1px solid #e5e7eb; border-radius:18px; box-shadow:0 5px 18px rgba(15,23,42,.08); padding:16px 18px 14px; margin-bottom:14px; }}
    .pas-upload-title {{ color:#0A0A0A; font-size:16px; font-weight:950; margin-bottom:10px; }}
    div[data-testid="stFileUploader"] {{ margin:0 !important; }}
    div[data-testid="stFileUploader"] label {{ display:none !important; }}
    div[data-testid="stFileUploader"] section {{ background:#f4f6f8 !important; border:1px solid #dfe4ea !important; border-radius:11px !important; min-height:52px !important; padding:8px 10px !important; }}
    div[data-testid="stFileUploader"] section * {{ color:#0A0A0A !important; }}
    div[data-testid="stFileUploader"] button {{ background:#fff !important; color:#0A0A0A !important; border:1px solid #d7dce3 !important; border-radius:10px !important; font-weight:900 !important; box-shadow:0 2px 8px rgba(0,0,0,.06) !important; }}
    div[data-testid="stFileUploader"] svg {{ color:#0A0A0A !important; fill:currentColor !important; stroke:currentColor !important; }}
    div[data-testid="stFileUploader"] [data-testid="stFileUploaderFile"] {{ background:#fff !important; border:1px solid #dfe4ea !important; border-radius:10px !important; color:#0A0A0A !important; }}
    div[data-testid="stFileUploader"] small {{ color:#4b5563 !important; }}
    div.stButton > button[kind="secondary"], .stButton > button {{ min-height:52px !important; font-size:16px !important; box-shadow:0 6px 18px rgba(255,212,0,.25) !important; }}
    .stDownloadButton > button {{ min-height:62px !important; font-size:20px !important; box-shadow:0 6px 18px rgba(255,212,0,.25) !important; }}

    .kpi-card {{ background:#fff !important; border-radius:18px !important; border:1px solid #e4e7eb !important; box-shadow:0 5px 20px rgba(15,23,42,.08) !important; min-height:118px !important; padding:18px 22px !important; display:flex; align-items:center; gap:18px; }}
    .kpi-icon {{ width:64px; height:64px; border-radius:50%; background:#fff5bd; display:flex; align-items:center; justify-content:center; flex:none; }}
    .kpi-icon svg {{ width:35px; height:35px; stroke:#0A0A0A; stroke-width:2.5; fill:none; stroke-linecap:round; stroke-linejoin:round; }}
    .kpi-label {{ color:#111 !important; font-size:15px !important; font-weight:950 !important; margin:0 0 3px !important; }}
    .kpi-value {{ color:#e9b900 !important; font-size:42px !important; line-height:.98 !important; font-weight:950 !important; text-shadow:none !important; }}
    .kpi-sub {{ color:#374151 !important; font-size:14px !important; margin-top:6px !important; }}

    .pas-results-title {{ color:#0A0A0A !important; font-size:28px !important; font-weight:950 !important; margin: 22px 0 8px !important; }}
    .pas-unmatched-pill {{ background:{PAS_YELLOW} !important; color:#0A0A0A !important; border:0 !important; border-radius:14px 14px 0 0 !important; padding:13px 20px !important; font-size:18px; box-shadow:0 4px 14px rgba(0,0,0,.09); }}
    .pas-table-wrap {{ background:#fff !important; border:1px solid #e0e4e9 !important; border-radius:0 16px 16px 16px !important; max-height:430px !important; overflow:auto !important; box-shadow:0 7px 25px rgba(15,23,42,.10) !important; }}
    table.pas-table {{ font-size:14px !important; color:#0A0A0A !important; }}
    table.pas-table thead th {{ background:{PAS_YELLOW} !important; color:#0A0A0A !important; border:1px solid #e2ba00 !important; padding:12px 14px !important; font-weight:950 !important; position:sticky; top:0; z-index:5; }}
    table.pas-table tbody td {{ background:#fff !important; color:#0A0A0A !important; border:1px solid #e1e5eb !important; padding:10px 14px !important; }}
    table.pas-table tbody tr:nth-child(even) td {{ background:#fbfcfd !important; }}

    div[data-testid="stFileUploader"] [data-testid="stFileUploaderFile"] {{ display: none !important; }}
    div[data-testid="stFileUploaderDropzone"] {{ background: transparent !important; border: 0 !important; padding: 0 !important; min-height: 0 !important; }}
    div[data-testid="stFileUploaderDropzoneInstructions"] {{ display: none !important; }}
    div[data-testid="stFileUploader"] section {{ background: transparent !important; border: 0 !important; min-height: 0 !important; padding: 0 !important; }}
    .pas-file-card {{ display:flex; align-items:center; gap:14px; background:#f4f6f8; border:1px solid #dfe4ea; border-radius:12px; padding:11px 14px; min-height:54px; margin: 4px 0 12px; }}
    .pas-file-icon {{ width:32px; height:32px; border-radius:8px; display:flex; align-items:center; justify-content:center; color:#fff; font-weight:950; font-size:11px; box-shadow:0 2px 8px rgba(0,0,0,.12); flex:none; }}
    .pas-file-icon.excel {{ background:#118a3b; }}
    .pas-file-main {{ flex:1; min-width:0; }}
    .pas-file-name {{ color:#0A0A0A; font-weight:950; font-size:15px; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }}
    .pas-file-size {{ color:#4b5563; font-weight:650; font-size:13px; margin-top:2px; }}
    .pas-file-check {{ width:24px; height:24px; border-radius:50%; background:#108a37; color:white; display:flex; align-items:center; justify-content:center; font-size:15px; font-weight:950; flex:none; }}
    </style>
    """,
    unsafe_allow_html=True,
)



st.markdown(
    """
    <style>
    /* Bottom chase animation: small, low, runs once */
    .pas-bottom-chase-wrap {
        position: fixed;
        left: calc(18rem + 22px);
        right: 42px;
        bottom: 12px;
        height: 58px;
        pointer-events: none;
        z-index: 1;
        overflow: hidden;
    }
    .pas-bottom-ground {
        position: absolute;
        left: 0;
        right: 0;
        bottom: 6px;
        border-bottom: 1px solid rgba(0,0,0,0.11);
    }
    .pas-chase-pack {
        position: absolute;
        bottom: 8px;
        left: -150px;
        width: 150px;
        height: 48px;
        animation: pas-chase-run 13s linear 1 forwards;
    }
    @keyframes pas-chase-run {
        0% { transform: translateX(-120px); opacity: 0; }
        8% { opacity: 1; }
        88% { opacity: 1; }
        100% { transform: translateX(calc(100vw - 90px)); opacity: 0; }
    }
    .pas-truck-mini { position: absolute; left: 0; bottom: 5px; width: 54px; height: 30px; filter: drop-shadow(0 1px 1px rgba(0,0,0,.22)); }
    .pas-truck-bed { position: absolute; left: 0; top: 5px; width: 34px; height: 19px; background: #FFD400; border: 3px solid #0A0A0A; border-radius: 4px 2px 3px 5px; transform: skewX(-10deg); }
    .pas-truck-logo { position: absolute; left: 7px; top: 9px; font-size: 9px; font-weight: 950; color: #0A0A0A; line-height: 1; z-index: 3; }
    .pas-truck-cab { position: absolute; left: 30px; top: 7px; width: 19px; height: 18px; background: #FFD400; border: 3px solid #0A0A0A; border-radius: 3px 5px 3px 2px; z-index: 2; }
    .pas-truck-window { position: absolute; left: 34px; top: 10px; width: 7px; height: 7px; background: #a8d8e8; border: 2px solid #0A0A0A; border-radius: 2px; z-index: 4; }
    .pas-truck-nose { position: absolute; left: 47px; top: 17px; width: 8px; height: 8px; background: #FFD400; border: 3px solid #0A0A0A; border-left: none; border-radius: 0 3px 3px 0; }
    .pas-wheel { position: absolute; bottom: 0; width: 9px; height: 9px; background: #0A0A0A; border: 2px solid #222; border-radius: 50%; animation: pas-wheel-spin .32s linear infinite; z-index: 5; }
    .pas-wheel::after { content: ""; position: absolute; inset: 2px; background: #FFD400; border-radius: 50%; }
    .pas-wheel.back { left: 13px; }
    .pas-wheel.front { left: 41px; }
    @keyframes pas-wheel-spin { to { transform: rotate(360deg); } }
    .pas-speed-lines { position: absolute; left: -30px; top: 17px; width: 24px; height: 18px; }
    .pas-speed-lines span { display:block; height:2px; background:#b9b9b9; margin:4px 0; border-radius:2px; animation: pas-flicker .55s linear infinite; }
    .pas-speed-lines span:nth-child(2) { width: 16px; margin-left: 8px; }
    .pas-speed-lines span:nth-child(3) { width: 11px; margin-left: 13px; }
    @keyframes pas-flicker { 50% { opacity:.25; transform: translateX(-5px); } }
    .pas-dust { position:absolute; left:-5px; bottom:0; width:34px; height:14px; opacity:.75; }
    .pas-dust span { position:absolute; bottom:0; background:#dac6a9; border-radius:50%; animation: pas-dust 1s linear infinite; }
    .pas-dust span:nth-child(1) { width:12px; height:6px; left:0; }
    .pas-dust span:nth-child(2) { width:16px; height:7px; left:10px; animation-delay:.2s; }
    .pas-dust span:nth-child(3) { width:11px; height:5px; left:23px; animation-delay:.4s; }
    @keyframes pas-dust { 50% { transform: translateX(-8px) scale(1.15); opacity:.4; } }
    .pas-stickman { position: absolute; left: 92px; bottom: 5px; width: 28px; height: 34px; animation: pas-runner-bob .35s ease-in-out infinite alternate; }
    @keyframes pas-runner-bob { from { transform: translateY(1px); } to { transform: translateY(-2px); } }
    .pas-stick-head { position:absolute; top:0; left:11px; width:8px; height:8px; border:2px solid #111; border-radius:50%; background:white; }
    .pas-stick-body { position:absolute; left:15px; top:9px; width:2px; height:13px; background:#111; transform: rotate(12deg); transform-origin:top; }
    .pas-stick-arm-a, .pas-stick-arm-b, .pas-stick-leg-a, .pas-stick-leg-b { position:absolute; width:2px; height:12px; background:#111; transform-origin:top; border-radius:2px; }
    .pas-stick-arm-a { left:15px; top:11px; transform: rotate(58deg); animation: pas-arm-a .35s linear infinite alternate; }
    .pas-stick-arm-b { left:15px; top:11px; transform: rotate(-50deg); animation: pas-arm-b .35s linear infinite alternate; }
    .pas-stick-leg-a { left:16px; top:21px; height:14px; transform: rotate(48deg); animation: pas-leg-a .35s linear infinite alternate; }
    .pas-stick-leg-b { left:16px; top:21px; height:14px; transform: rotate(-42deg); animation: pas-leg-b .35s linear infinite alternate; }
    @keyframes pas-arm-a { to { transform: rotate(-45deg); } }
    @keyframes pas-arm-b { to { transform: rotate(55deg); } }
    @keyframes pas-leg-a { to { transform: rotate(-45deg); } }
    @keyframes pas-leg-b { to { transform: rotate(48deg); } }
    </style>
    """,
    unsafe_allow_html=True,
)


def render_bottom_chase():
    st.markdown(
        """
        <div class="pas-bottom-chase-wrap" aria-hidden="true">
            <div class="pas-bottom-ground"></div>
            <div class="pas-chase-pack">
                <div class="pas-speed-lines"><span></span><span></span><span></span></div>
                <div class="pas-dust"><span></span><span></span><span></span></div>
                <div class="pas-truck-mini">
                    <div class="pas-truck-bed"></div>
                    <div class="pas-truck-logo">PAS</div>
                    <div class="pas-truck-cab"></div>
                    <div class="pas-truck-window"></div>
                    <div class="pas-truck-nose"></div>
                    <div class="pas-wheel back"></div>
                    <div class="pas-wheel front"></div>
                </div>
                <div class="pas-stickman">
                    <div class="pas-stick-head"></div>
                    <div class="pas-stick-body"></div>
                    <div class="pas-stick-arm-a"></div>
                    <div class="pas-stick-arm-b"></div>
                    <div class="pas-stick-leg-a"></div>
                    <div class="pas-stick-leg-b"></div>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def logo_available() -> bool:
    return Path("pas_logo.png").exists() or Path("assets/pas_logo.png").exists()

with st.sidebar:
    if Path("pas_logo.png").exists():
        st.image("pas_logo.png", use_column_width=True)
    elif Path("assets/pas_logo.png").exists():
        st.image("assets/pas_logo.png", use_column_width=True)
    else:
        st.markdown('<div style="background:#FFD400;color:#000;border-radius:14px;padding:18px;text-align:center;font-weight:950;font-size:30px;">PAS</div>', unsafe_allow_html=True)
    st.markdown(
        """
        <div class="pas-sidebar-title">PAS Hire<br>Report Builder</div>
        <div class="pas-yellow-line"></div>
        <div class="pas-sidebar-copy">Upload the Material & Plant Orders workbook, then export completed Excel and site PDFs.</div>
        <div class="pas-sidebar-rule"></div>
        <div class="pas-sidebar-heading">Instructions</div>
        <div class="pas-nav-row"><span class="pas-nav-icon"><svg viewBox="0 0 24 24"><path d="M16 16l-4-4-4 4"/><path d="M12 12v9"/><path d="M20 16.6A5 5 0 0 0 18 7h-1.3A8 8 0 1 0 4 15.3"/></svg></span><span>Upload Material & Plant Orders</span></div>
        <div class="pas-nav-row"><span class="pas-nav-icon"><svg viewBox="0 0 24 24"><path d="M5 3l14 9-14 9V3z"/></svg></span><span>Build Hire Reports</span></div>
        <div class="pas-nav-row"><span class="pas-nav-icon"><svg viewBox="0 0 24 24"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><path d="M7 10l5 5 5-5"/><path d="M12 15V3"/></svg></span><span>Download Excel</span></div>
        <div class="pas-nav-row"><span class="pas-nav-icon"><svg viewBox="0 0 24 24"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><path d="M7 10l5 5 5-5"/><path d="M12 15V3"/></svg></span><span>Download PDF Pack</span></div>
        <div class="pas-nav-row"><span class="pas-nav-icon"><svg viewBox="0 0 24 24"><circle cx="11" cy="11" r="8"/><path d="M21 21l-4.3-4.3"/></svg></span><span>Smoke Crack</span></div>
        <div class="pas-sidebar-rule"></div>
        <div class="pas-sidebar-footer">PAS NW Ltd • v1.0 Prototype Build</div>
        """,
        unsafe_allow_html=True,
    )

st.markdown(
    f"""
    <div class="pas-hero">
      <div class="pas-hero-logo">PAS</div>
      <div class="pas-hero-text">PAS NW Ltd<span class="pas-hero-dot">•</span><span class="pas-hero-version">{APP_VERSION}</span></div>
    </div>
    """,
    unsafe_allow_html=True,
)


def render_selected_file_card(uploaded_file, file_kind="excel"):
    size = getattr(uploaded_file, "size", 0) or 0
    if size >= 1024 * 1024:
        size_text = f"{size / (1024 * 1024):.1f} MB"
    else:
        size_text = f"{size / 1024:.0f} KB"
    st.markdown(
        f'''
        <div class="pas-file-card">
            <div class="pas-file-icon excel">XLS</div>
            <div class="pas-file-main">
                <div class="pas-file-name">{escape(getattr(uploaded_file, "name", "Uploaded file"))}</div>
                <div class="pas-file-size">{size_text}</div>
            </div>
            <div class="pas-file-check">✓</div>
        </div>
        ''',
        unsafe_allow_html=True,
    )


def clean_cell(value) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    text = str(value).strip()
    if text.lower() in {"nan", "none", "nat"}:
        return ""
    return text


def norm_header(value) -> str:
    return re.sub(r"[^a-z0-9]+", "", clean_cell(value).lower())


def find_required_columns(ws):
    headers = {norm_header(cell.value): cell.column for cell in ws[1]}
    aliases = {
        "Description": ["description"],
        "Fleet No.": ["fleetno", "fleetnumber", "fleet"],
        "Supplier": ["supplier"],
        "Qty": ["qty", "quantity"],
        "On Hire / Delivery Date": ["onhiredeliverydate", "onhiredate", "deliverydate"],
        "Expected Off-Hire Date": ["expectedoffhiredate", "expectedoffhire"],
        "Off Hire Date": ["offhiredate", "offhire"],
        "Status": ["status"],
        "Job No": ["jobno", "jobnumber", "jobnr"],
        "Site Name": ["sitename", "site"],
        "User/Gang": ["usergang", "user", "gang"],
    }
    found = {}
    missing = []
    for logical, keys in aliases.items():
        col = None
        for key in keys:
            if key in headers:
                col = headers[key]
                break
        if col is None:
            missing.append(logical)
        found[logical] = col
    if missing:
        raise ValueError("Could not find required Plant columns: " + ", ".join(missing))
    return found


def format_date(value):
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    return clean_cell(value)


def safe_sheet_name(name: str) -> str:
    text = clean_cell(name) or "Unknown"
    text = re.sub(r"[\\/*?:\[\]]", "-", text).strip()
    return text[:31] or "Unknown"


def safe_filename(name: str) -> str:
    text = clean_cell(name) or "Unknown"
    return re.sub(r"[^A-Za-z0-9_. -]+", "-", text).strip() or "Unknown"


def locate_template_path() -> Path:
    for path in TEMPLATE_PATHS:
        if path.exists():
            return path
    raise FileNotFoundError("Built-in template not found. Put 'Hire Report Template.xlsx' in the assets folder.")


def copy_row_style(ws, source_row: int, target_row: int, max_col: int):
    for col in range(1, max_col + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.border:
            dst.border = copy(src.border)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.font:
            dst.font = copy(src.font)
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def clear_body_rows(ws, start_row=2, max_col=9):
    for row in range(start_row, ws.max_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row, col).value = None


def extract_hire_rows(uploaded_file) -> pd.DataFrame:
    source_wb = load_workbook(uploaded_file, data_only=True)
    if "Plant" not in source_wb.sheetnames:
        raise ValueError("Could not find a 'Plant' tab in the uploaded workbook.")
    ws = source_wb["Plant"]
    cols = find_required_columns(ws)
    rows = []
    for r in range(2, ws.max_row + 1):
        status = clean_cell(ws.cell(r, cols["Status"]).value)
        if status.lower().strip() not in {"on hire", "missing"}:
            continue
        job_no = clean_cell(ws.cell(r, cols["Job No"]).value)
        if not job_no:
            continue
        rows.append({
            "Description": ws.cell(r, cols["Description"]).value,
            "Fleet No.": ws.cell(r, cols["Fleet No."]).value,
            "Supplier": ws.cell(r, cols["Supplier"]).value,
            "Qty": ws.cell(r, cols["Qty"]).value,
            "On Hire / Delivery Date": ws.cell(r, cols["On Hire / Delivery Date"]).value,
            "Expected Off-Hire Date": ws.cell(r, cols["Expected Off-Hire Date"]).value,
            "Off Hire Date": ws.cell(r, cols["Off Hire Date"]).value,
            "Job No": job_no,
            "User/Gang": ws.cell(r, cols["User/Gang"]).value,
            "Site Name": ws.cell(r, cols["Site Name"]).value,
            "Status": status,
        })
    df = pd.DataFrame(rows)
    if df.empty:
        return df
    df["Supplier_sort"] = df["Supplier"].map(lambda x: clean_cell(x).lower())
    df["Description_sort"] = df["Description"].map(lambda x: clean_cell(x).lower())
    df = df.sort_values(["Job No", "Supplier_sort", "Description_sort"]).drop(columns=["Supplier_sort", "Description_sort"])
    return df


def write_hire_audit(wb, grouped, run_date):
    if "Hire Audit" not in wb.sheetnames:
        raise ValueError("Template is missing the 'Hire Audit' tab.")
    ws = wb["Hire Audit"]

    # Clear old values but keep the sheet structure.
    for row in range(2, ws.max_row + 1):
        for col in range(1, 14):
            ws.cell(row, col).value = None

    # Force consistent borders/styles from the known good template row.
    # This fixes the occasional missing grid lines, e.g. rows 20/21 on Hire Audit.
    style_source_row = 2

    row_num = 2
    for job_no, data in grouped:
        site_name = clean_cell(data["Site Name"].iloc[0]) or ""
        copy_row_style(ws, style_source_row, row_num, 13)

        ws.cell(row_num, 1).value = job_no
        ws.cell(row_num, 2).value = site_name
        ws.cell(row_num, 3).value = run_date
        ws.cell(row_num, 3).number_format = "dd/mm/yyyy"

        # Only A, B and C are completed. Everything else stays blank.
        for col in range(4, 14):
            ws.cell(row_num, col).value = None
        row_num += 1

    # Also normalise the visible blank rows so the printed sheet never has missing borders.
    for blank_row in range(row_num, ws.max_row + 1):
        copy_row_style(ws, style_source_row, blank_row, 13)
        for col in range(1, 14):
            ws.cell(blank_row, col).value = None

def build_excel_workbook(template_path: Path, hire_df: pd.DataFrame):
    wb = load_workbook(template_path)
    if "Site No" not in wb.sheetnames:
        raise ValueError("Template is missing the 'Site No' tab.")
    template_ws = wb["Site No"]
    run_date = datetime.now().date()
    grouped = list(hire_df.groupby("Job No", sort=True))
    output_columns = [
        "Description", "Fleet No.", "Supplier", "Qty", "On Hire / Delivery Date",
        "Expected Off-Hire Date", "Off Hire Date", "Job No", "User/Gang"
    ]
    created = []
    for job_no, data in grouped:
        if data.empty:
            continue
        ws = wb.copy_worksheet(template_ws)
        ws.title = safe_sheet_name(job_no)
        clear_body_rows(ws, 2, 9)
        data = data.sort_values(["Supplier", "Description"], key=lambda s: s.map(lambda x: clean_cell(x).lower()))
        for idx, (_, record) in enumerate(data.iterrows(), start=2):
            if idx > ws.max_row:
                copy_row_style(ws, 2, idx, 9)
            for col_idx, col_name in enumerate(output_columns, start=1):
                cell = ws.cell(idx, col_idx)
                value = record.get(col_name, "")
                cell.value = value
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                if col_name in {"On Hire / Delivery Date", "Expected Off-Hire Date", "Off Hire Date"}:
                    cell.number_format = "dd/mm/yyyy"
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
        created.append({
            "Job No": job_no,
            "Site Name": clean_cell(data["Site Name"].iloc[0]),
            "Rows": len(data),
            "Worksheet": ws.title,
            "Data": data,
        })
    write_hire_audit(wb, grouped, run_date)
    if "Site No" in wb.sheetnames:
        del wb["Site No"]
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), created, run_date


def make_pdf_for_job(job_no: str, site_name: str, data: pd.DataFrame, run_date) -> bytes:
    if colors is None:
        raise RuntimeError("ReportLab is unavailable. Add reportlab to requirements.txt.")
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=8 * mm,
        leftMargin=8 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("PAS_Title", parent=styles["Heading1"], fontName="Helvetica-Bold", fontSize=13, leading=15, textColor=colors.black, spaceAfter=2)
    sub_style = ParagraphStyle("PAS_Sub", parent=styles["Normal"], fontName="Helvetica", fontSize=8, leading=10, textColor=colors.black)
    cell_style = ParagraphStyle("PAS_Cell", parent=styles["Normal"], fontName="Helvetica", fontSize=6.4, leading=7.4, textColor=colors.black)
    head_style = ParagraphStyle("PAS_Head", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=6.4, leading=7.4, textColor=colors.black)
    story = []
    story.append(Paragraph(f"Hire Report - {escape(clean_cell(job_no))}", title_style))
    if site_name:
        story.append(Paragraph(escape(clean_cell(site_name)), sub_style))
    story.append(Paragraph(f"Generated: {run_date.strftime('%d/%m/%Y')}", sub_style))
    story.append(Spacer(1, 4))
    headers = ["Description", "Fleet No.", "Supplier", "Qty", "On Hire / Delivery Date", "Expected Off-Hire Date", "Off Hire Date", "Job No", "User/Gang"]
    table_data = [[Paragraph(h, head_style) for h in headers]]
    data = data.sort_values(["Supplier", "Description"], key=lambda s: s.map(lambda x: clean_cell(x).lower()))
    for _, row in data.iterrows():
        table_data.append([
            Paragraph(escape(clean_cell(row.get("Description", ""))), cell_style),
            Paragraph(escape(clean_cell(row.get("Fleet No.", ""))), cell_style),
            Paragraph(escape(clean_cell(row.get("Supplier", ""))), cell_style),
            Paragraph(escape(clean_cell(row.get("Qty", ""))), cell_style),
            Paragraph(escape(format_date(row.get("On Hire / Delivery Date", ""))), cell_style),
            Paragraph(escape(format_date(row.get("Expected Off-Hire Date", ""))), cell_style),
            Paragraph(escape(format_date(row.get("Off Hire Date", ""))), cell_style),
            Paragraph(escape(clean_cell(row.get("Job No", ""))), cell_style),
            Paragraph(escape(clean_cell(row.get("User/Gang", ""))), cell_style),
        ])
    col_widths = [54*mm, 20*mm, 32*mm, 12*mm, 27*mm, 29*mm, 24*mm, 18*mm, 44*mm]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(PAS_YELLOW)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#8c8c8c")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(table)
    doc.build(story)
    return buffer.getvalue()


def build_pdf_zip(created_jobs, run_date) -> bytes:
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for item in created_jobs:
            pdf_bytes = make_pdf_for_job(item["Job No"], item["Site Name"], item["Data"], run_date)
            zf.writestr(f"{safe_filename(item['Job No'])}.pdf", pdf_bytes)
    return zip_buffer.getvalue()


def render_jobs_table(created_jobs):
    if not created_jobs:
        st.markdown('<div class="pas-unmatched-pill">Generated Reports</div>', unsafe_allow_html=True)
        st.markdown('<div class="pas-table-wrap"><table class="pas-table"><tbody><tr><td>No reports generated.</td></tr></tbody></table></div>', unsafe_allow_html=True)
        return
    headers = ["Job No", "Site Name", "Rows", "PDF"]
    header_html = "".join(f"<th>{escape(h)}</th>" for h in headers)
    rows_html = []
    for item in created_jobs:
        rows_html.append(
            "<tr>"
            f"<td>{escape(clean_cell(item['Job No']))}</td>"
            f"<td>{escape(clean_cell(item['Site Name']))}</td>"
            f"<td>{item['Rows']}</td>"
            f"<td>{escape(safe_filename(item['Job No']))}.pdf</td>"
            "</tr>"
        )
    st.markdown('<div class="pas-unmatched-pill">Generated Reports</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="pas-table-wrap"><table class="pas-table"><thead><tr>{header_html}</tr></thead><tbody>{"".join(rows_html)}</tbody></table></div>', unsafe_allow_html=True)


st.markdown('<div class="pas-upload-card"><div class="pas-upload-title">Upload Material & Plant Orders workbook</div>', unsafe_allow_html=True)
orders_file = st.file_uploader("Upload Material & Plant Orders", type=["xlsx", "xlsm", "xls"], label_visibility="collapsed", key="orders_upload")
if orders_file:
    st.markdown(
        """
        <style>
        /* Hide the Streamlit uploader controls once a file is selected. */
        div[data-testid="stFileUploader"] {
            display: none !important;
            visibility: hidden !important;
            height: 0 !important;
            min-height: 0 !important;
            margin: 0 !important;
            padding: 0 !important;
            overflow: hidden !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )
    render_selected_file_card(orders_file, "excel")
st.markdown('</div>', unsafe_allow_html=True)

run = st.button("▶  Build hire reports", use_container_width=True)

if "hire_report_results" not in st.session_state:
    st.session_state["hire_report_results"] = None

if run:
    if not orders_file:
        st.warning("Please upload the Material & Plant Orders workbook.")
        st.stop()
    try:
        template_path = locate_template_path()
        with st.spinner("Reading Plant tab..."):
            hire_df = extract_hire_rows(orders_file)
        if hire_df.empty:
            st.warning("No On Hire or Missing plant rows were found.")
            st.stop()
        with st.spinner("Building completed hire report workbook..."):
            excel_bytes, created_jobs, run_date = build_excel_workbook(template_path, hire_df)
        with st.spinner("Creating PDF pack..."):
            pdf_zip_bytes = build_pdf_zip(created_jobs, run_date)
        stamp = datetime.now().strftime("%Y%m%d_%H%M")
        st.session_state["hire_report_results"] = {
            "hire_df": hire_df,
            "created_jobs": created_jobs,
            "run_date": run_date,
            "excel_bytes": excel_bytes,
            "pdf_zip_bytes": pdf_zip_bytes,
            "total_rows": len(hire_df),
            "total_sites": len(created_jobs),
            "excel_filename": f"PAS_Hire_Report_Audit_{stamp}.xlsx",
            "pdf_zip_filename": f"PAS_Hire_Report_PDFs_{stamp}.zip",
        }
    except Exception as e:
        st.error(f"Something went wrong: {e}")
        st.exception(e)

results = st.session_state.get("hire_report_results")

if results is not None:
    total_rows = results["total_rows"]
    total_sites = results["total_sites"]
    run_date = results["run_date"].strftime("%d/%m/%Y")
    pdf_count = len(results["created_jobs"])

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.markdown(f'<div class="kpi-card"><div class="kpi-icon"><svg viewBox="0 0 24 24"><path d="M8 7V3h8l4 4v14H6V7z"/><path d="M16 3v5h5"/><path d="M9 13h6"/><path d="M9 17h4"/><path d="M4 7h2v14h12"/></svg></div><div><div class="kpi-label">Hire rows</div><div class="kpi-value">{total_rows}</div><div class="kpi-sub">On Hire + Missing</div></div></div>', unsafe_allow_html=True)
    with c2:
        st.markdown(f'<div class="kpi-card"><div class="kpi-icon"><svg viewBox="0 0 24 24"><path d="M3 21h18"/><path d="M5 21V7l8-4v18"/><path d="M19 21V11l-6-4"/><path d="M9 9h.01"/><path d="M9 13h.01"/><path d="M9 17h.01"/></svg></div><div><div class="kpi-label">Sites</div><div class="kpi-value">{total_sites}</div><div class="kpi-sub">Job tabs created</div></div></div>', unsafe_allow_html=True)
    with c3:
        st.markdown(f'<div class="kpi-card"><div class="kpi-icon"><svg viewBox="0 0 24 24"><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/><path d="M14 2v6h6"/><path d="M9 13h6"/><path d="M9 17h6"/></svg></div><div><div class="kpi-label">PDFs</div><div class="kpi-value">{pdf_count}</div><div class="kpi-sub">One per site</div></div></div>', unsafe_allow_html=True)
    with c4:
        st.markdown(f'<div class="kpi-card"><div class="kpi-icon"><svg viewBox="0 0 24 24"><path d="M8 2v4"/><path d="M16 2v4"/><rect x="3" y="4" width="18" height="18" rx="2"/><path d="M3 10h18"/></svg></div><div><div class="kpi-label">Run date</div><div class="kpi-value" style="font-size:30px;">{run_date}</div><div class="kpi-sub">Hire Audit date sent</div></div></div>', unsafe_allow_html=True)

    st.markdown('<div class="pas-results-title">Results</div>', unsafe_allow_html=True)
    render_jobs_table(results["created_jobs"])

    dl_left, dl_right = st.columns(2)
    with dl_left:
        st.download_button(
            "⬇  Download Excel",
            data=results["excel_bytes"],
            file_name=results["excel_filename"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with dl_right:
        st.download_button(
            "⬇  Download PDF ZIP",
            data=results["pdf_zip_bytes"],
            file_name=results["pdf_zip_filename"],
            mime="application/zip",
            use_container_width=True,
        )
else:
    st.info("Upload Material & Plant Orders, then click Build hire reports.")


render_bottom_chase()
